
import pywhatkit as kit
import time
from datetime import datetime
from openpyxl import load_workbook

#Ver as abas
#print(arquivo.sheetnames)

#Chamando o arquivo excel
arquivo = load_workbook("lista_contatos_ficticios.xlsx")
#ver as abas
aba_nomes = arquivo["Contatos"]
#Quantidade de linhas no excel
linhas = aba_nomes.max_row

nome = aba_nomes.cell(row=1, column=1).value
print(nome)

for row in aba_nomes.iter_rows(min_row=2, values_only=True):
        nome, telefone, mensagem = row
        
        if not telefone or not mensagem:
            continue

        print(f"Enviando mensagem para {nome} ({telefone})...")
        
        try:
            # pywhatkit.sendwhatmsg_instantly
            # Esoerar 15 seundos: tempo de espera antes de enviar (segundos)
            # tab_close: se deve fechar a aba após enviar
            kit.sendwhatmsg_instantly(
                phone_no=str(telefone),
                message=mensagem,
                wait_time=15,
                tab_close=True,
                close_time=3
            )
            # Mostra se a mensagem foi enviada com sucesso
            print(f"Sucesso ao agendar envio para {nome}!")
            
            # Pequena pausa entre os envios para evitar bloqueios
            time.sleep(5)
            
        except Exception as e:
            # Mostra se a mensagem não foi enviada  
            print(f"Erro ao enviar para {nome}: {e}")
    

#para salvar o arquivo caso tenha alguma alteração no sistema
#arquivo.save("lista_contatos_ficticios.xlsx")
